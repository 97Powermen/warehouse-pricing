"""
一键重新生成海外仓场景收费展示页：
1) 从 Excel 源文件提取 3 个区域（US/EU/UK）的 6 场景数据
2) 提取「收费说明」sheet 的说明文字与表格
3) 注入 template.html 模板
4) 输出 warehouse_pricing_dashboard.html 与 index.html（云端部署入口）

用法：
  python build_dashboard.py
（可改下面的 EXCEL_PATH 指向你的源文件）
"""
import json
from openpyxl import load_workbook

EXCEL_PATH = r'F:/桌面/PNP与NJNY仓场景收费明细表.xlsx'
TEMPLATE = r'C:/Users/EJET/WorkBuddy/2026-08-05-18-36-39/outputs/template.html'
OUTPUT = r'C:/Users/EJET/WorkBuddy/2026-08-05-18-36-39/outputs/warehouse_pricing_dashboard.html'
# 部署到云端用的入口文件名（与 OUTPUT 内容完全一致，方便一键更新）
INDEX = r'C:/Users/EJET/WorkBuddy/2026-08-05-18-36-39/outputs/index.html'

# 6 个标准场景（顺序必须与各 sheet 的列顺序一致）
SCENARIOS = [
    '只更换Fnsku标',
    '更换防伪标+FNSKU标',
    '更换防伪标+FNSKU标品牌标',
    '只更换Fnsku标+升级',
    '更换防伪标+FNSKU标+升级',
    '更换防伪标+FNSKU标+品牌标+升级',
]

# 各区域 sheet 的列布局：每个仓库有 套装/单支 两套起始列，每套 6 个场景连续排列
SHEET_CONFIG = {
    'US': {
        'sheet': 'US区域',
        'name': 'US区域（美仓）',
        'subtitle': 'PNP仓 vs NJ/NY仓 · 双仓对比',
        'compare': True,
        'region_label': 'US',
        'warehouses': {
            'PNP': {'set': 60, 'single': 72},       # BH-BM / BT-BY
            'NJ/NY': {'set': 66, 'single': 78},     # BN-BS / BZ-CE
        },
    },
    'EU': {
        'sheet': 'EU区域',
        'name': 'EU区域（德仓）',
        'subtitle': '德仓（DEC）· 单仓展示',
        'compare': False,
        'region_label': 'EU',
        'warehouses': {
            'DEC': {'set': 40, 'single': 46},       # AN-AS / AT-AY
        },
    },
    'UK': {
        'sheet': 'UK区域',
        'name': 'UK区域（英仓）',
        'subtitle': '英仓（UK）· 单仓展示',
        'compare': False,
        'region_label': 'UK',
        'warehouses': {
            'UK': {'set': 39, 'single': 45},        # AM-AR / AS-AX
        },
    },
}


def safe_float(v):
    """数字正常转换；空白或非数字一律视为无此场景 -> None"""
    if v is None:
        return None
    try:
        return round(float(v), 4)
    except (ValueError, TypeError):
        return None


def is_business_missing(v):
    """某些升级业务暂未开展，单元格会填写类似“无单支升级场景”的文字"""
    if v is None:
        return False
    return isinstance(v, str) and ('无' in v or '未' in v or '暂无' in v or '/' == v.strip())


def extract_region(cfg):
    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb[cfg['sheet']]
    products = []
    rid = 0
    for row_idx in range(4, ws.max_row + 1):
        product = ws.cell(row=row_idx, column=1).value
        if not product:
            continue
        raw_up = ws.cell(row=row_idx, column=3).value
        is_up = str(raw_up).strip() if raw_up else '否'
        p_count = ws.cell(row=row_idx, column=4).value
        raw_region = ws.cell(row=row_idx, column=2).value
        region = str(raw_region).strip() if raw_region else cfg['region_label']
        scn = {}
        for name in SCENARIOS:
            scn[name] = {}
            idx = SCENARIOS.index(name)
            for wh, cols in cfg['warehouses'].items():
                scn[name][wh] = {}
                for unit in ('set', 'single'):
                    col = cols[unit] + idx
                    raw = ws.cell(row=row_idx, column=col).value
                    if unit == 'single' and is_business_missing(raw):
                        val = None
                    else:
                        val = safe_float(raw)
                    scn[name][wh][unit] = val
        is_upgrade = (is_up == '是')
        products.append({
            'id': rid,
            'product': str(product).strip(),
            'is_upgrade': is_upgrade,
            'is_upgrade_text': '易升级' if is_upgrade else '非易升级',
            'p_count': int(p_count) if p_count is not None and str(p_count).strip().isdigit() else None,
            'region': region,
            'scenarios': scn,
        })
        rid += 1
    return products


def symbol_for_region(region):
    region = str(region).strip().upper()
    if region == 'EU':
        return '€'
    if region == 'UK':
        return '£'
    return '$'


def format_fee_value(v, symbol):
    """给数字值追加货币符号；非数字原样返回"""
    if v is None:
        return ''
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return f"{symbol}{v}"
    return v


def extract_fee_note():
    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb['收费说明']
    label_text = ws['A1'].value
    label_headers = [ws.cell(row=2, column=c).value for c in range(1, 9)]
    label_rows = []
    for r in (3, 4, 5):
        raw = [ws.cell(row=r, column=c).value for c in range(1, 9)]
        region = str(raw[0]).strip() if raw[0] else ''
        symbol = symbol_for_region(region)
        label_rows.append([raw[0]] + [format_fee_value(v, symbol) for v in raw[1:]])

    upgrade_text = ws['A8'].value
    upgrade_headers = [ws.cell(row=9, column=c).value for c in range(1, 10)]
    # 升级费用各列所属区域：根据表头文字判断（德仓/EU -> €，英仓/UK -> £，其余 -> $）
    upgrade_col_symbols = []
    for c in range(2, 10):
        h = str(ws.cell(row=9, column=c).value or '').upper()
        if '德' in h or 'EU' in h:
            upgrade_col_symbols.append('€')
        elif '英' in h or 'UK' in h:
            upgrade_col_symbols.append('£')
        else:
            upgrade_col_symbols.append('$')
    upgrade_rows = []
    for r in (10, 11):
        raw = [ws.cell(row=r, column=c).value for c in range(1, 10)]
        upgrade_rows.append([raw[0]] + [format_fee_value(v, upgrade_col_symbols[i]) for i, v in enumerate(raw[1:])])

    return {
        'labelText': label_text,
        'labelHeaders': label_headers,
        'labelRows': label_rows,
        'upgradeText': upgrade_text,
        'upgradeHeaders': upgrade_headers,
        'upgradeRows': upgrade_rows,
    }


def build():
    regions = {}
    for key, cfg in SHEET_CONFIG.items():
        regions[key] = {
            'name': cfg['name'],
            'subtitle': cfg['subtitle'],
            'compare': cfg['compare'],
            'warehouses': list(cfg['warehouses'].keys()),
            'products': extract_region(cfg),
        }
    fee = extract_fee_note()
    data = {'regions': regions, 'feeNote': fee}

    with open(TEMPLATE, encoding='utf-8') as f:
        tpl = f.read()
    out = tpl.replace('__DATA__', json.dumps(data, ensure_ascii=False))

    with open(OUTPUT, 'w', encoding='utf-8') as f:
        f.write(out)
    with open(INDEX, 'w', encoding='utf-8') as f:
        f.write(out)
    with open(r'C:/Users/EJET/WorkBuddy/2026-08-05-18-36-39/outputs/warehouse_data.json', 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    total = sum(len(r['products']) for r in regions.values())
    print(f'✅ 已生成：{OUTPUT}')
    for k, r in regions.items():
        print(f'   {k} {r["name"]}: {len(r["products"])} 个产品，仓库 {r["warehouses"]}')
    print(f'   合计 {total} 个产品记录')


if __name__ == '__main__':
    build()
